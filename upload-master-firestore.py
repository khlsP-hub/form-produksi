#!/usr/bin/env python3
"""
upload-master-firestore.py
--------------------------
Upload data master produk, mesin, dan karyawan (karu & asisten) dari Excel ke Firestore.
Mendukung: PET, INJECT, BLOW, DECORATING, SECOND_PROSES.

Requirements:
  pip install firebase-admin pandas openpyxl

CARA PAKAI:
  # Upload semua sekaligus (rekomendasi):
  py upload-master-firestore.py serviceAccountKey.json

  # Simulasi dulu (tidak upload ke Firestore):
  py upload-master-firestore.py serviceAccountKey.json --dry-run

  # Hapus data lama dulu sebelum upload:
  py upload-master-firestore.py serviceAccountKey.json --clear
"""

import sys
import os
import argparse
import firebase_admin
from firebase_admin import credentials, firestore

# ═══════════════════════════════════════════════════════════════════
#  KONFIGURASI PRODUK & MESIN
# ═══════════════════════════════════════════════════════════════════

DATABASES = [
    {
        'bagian': 'PET',
        'file':   'Database produk PET upd.xlsx',
        'produk_sheets': [
            {'sheet': 'Database',    'col_kode': 'KODE BARANG', 'col_nama': 'PRODUK'},
            {'sheet': 'Database SB', 'col_kode': 'KODE BARANG', 'col_nama': 'PRODUK'},
        ],
        'mesin_sheets': [
            {'sheet': 'Database',    'col_mesin': 'NO MESIN'},
            {'sheet': 'Database SB', 'col_mesin': 'NO MESIN'},
        ],
    },
    {
        'bagian': 'INJECT',
        'file':   'Inject 08 April 2026.xlsx',
        'produk_sheets': [
            {'sheet': 'DATABASE', 'col_kode': 'KODE BARANG', 'col_nama': 'NAMA PRODUK'},
        ],
        'mesin_sheets': [
            {'sheet': 'DATABASE', 'col_mesin': 'Nomor Mesin'},
        ],
    },
    # {
    #     'bagian': 'BLOW',
    #     'file':   'Data Blow.xlsx',
    #     'produk_sheets': [
    #         {'sheet': 'Database', 'col_kode': 'KODE BARANG', 'col_nama': 'NAMA PRODUK'},
    #     ],
    #     'mesin_sheets': [
    #         {'sheet': 'Database', 'col_mesin': 'Nomor Mesin'},
    #     ],
    # },
]

# ═══════════════════════════════════════════════════════════════════
#  KONFIGURASI KARYAWAN
#
#  Tiga tipe format Excel yang didukung:
#
#  TYPE A — "inject"  : Header baris N, kolom per karu (horizontal)
#    Contoh: INJECT
#    Baris header: | No Mesin | Karu Febriyanto | Karu Hermanto | ...
#    Baris data:   | I-01     | Aep Saefulloh   | Tugiman       | ...
#
#  TYPE B — "pet"     : Kolom Karu (A) + Kolom Asisten (B), vertikal
#    Contoh: PET
#    Header: | KETUA REGU | NAMA ASISTEN | NO MC |
#    Data:   | RIAN       | SETIYONO     | P-02  |
#            | =+A4       | HARYANTO     | P-11  |  ← formula = karu sama
#
#  TYPE C — "regu"    : Header baris N, kolom per regu (horizontal), tanpa no mesin
#    Contoh: BLOW, DECORATING
#    Header: | NO | REGU SUGENG | REGU IBNU ZAMZAM | REGU SUGIYANTO |
#    Data:   | 1  | RURI        | BERNAT           | SAMBUDI        |
#    → Nama karu diambil dari header kolom (hapus prefix "REGU ")
#
#  TYPE D — "jabatan" : Header baris N, kolom per regu + kolom JABATAN
#    Contoh: SECOND_PROSES
#    Header: | NO | REGU ROHALI | REGU AAN | REGU IFAN | JABATAN |
#    Data:   | 1  | rohali      | aan      | Ifan      | KARU    |
#            | 2  | Ari         | Sokhib   | sacep     | ASSTN KARU |
#    → Baris dengan JABATAN=KARU → nama karu; sisanya → asisten
#
# ═══════════════════════════════════════════════════════════════════

KARYAWAN_DATABASES = [

    # ── INJECT (TYPE A) ───────────────────────────────────────────
    # {
    #     'bagian': 'INJECT',
    #     'file':   'Data Nama Asisten Departemen Injection.xlsx',
    #     'sheets': [
    #         {
    #             'type':         'inject',
    #             'sheet':        'Nama Asisten ',   # perhatikan spasi di nama sheet
    #             'header_row':   3,                 # baris header (1-based)
    #             'col_mesin':    2,                 # kolom No Mesin (1-based)
    #             'cols_asisten': [3, 4, 5],         # kolom-kolom nama asisten (1-based)
    #             'karu_prefix':  'Karu ',           # prefix di header → dihapus jadi nama karu
    #         },
    #     ],
    # },

    # ── PET (TYPE B) ──────────────────────────────────────────────
    # {
    #     'bagian': 'PET',
    #     'file':   'Pembagian Mesin Asisten PET.xlsx',
    #     'sheets': [
    #         {
    #             'type':       'pet',
    #             'sheet':      'Sheet1',
    #             'header_row': 3,   # baris header (1-based): KETUA REGU | NAMA ASISTEN | NO MC
    #             # col A (1) = karu, col B (2) = asisten, col C (3) = no mesin
    #         },
    #     ],
    # },

    # ── BLOW (TYPE C) ─────────────────────────────────────────────
    {
        'bagian': 'BLOW',
        'file':   'Data Nama Asisten Blow.xlsx',
        'sheets': [
            {
                'type':        'regu',
                'sheet':       'Sheet1',
                'header_row':  6,            # baris header: NO | REGU SUGENG | ...
                'col_no':      1,            # kolom nomor urut (skip jika kosong)
                'cols_regu':   [2, 3, 4],    # kolom-kolom per regu (1-based)
                'karu_prefix': 'REGU ',      # prefix di header → dihapus jadi nama karu
            },
        ],
    },

    # ── DECORATING (TYPE C) ───────────────────────────────────────
    {
        'bagian': 'DECORATING',
        'file':   'Data_nama_karru___asissten_karru_departemen_decorating.xlsx',
        'sheets': [
            {
                'type':        'regu',
                'sheet':       'Sheet1',
                'header_row':  6,            # baris header: NO | WARISUN | UJANG... | DEDE...
                'col_no':      1,
                'cols_regu':   [2, 3, 4],
                'karu_prefix': '',           # tidak ada prefix → baris 1 langsung jadi karu
                'karu_row':    1,            # nama di header_row = nama karu (baris pertama data = asisten)
            },
        ],
    },

    # ── SECOND PROSES (TYPE D) ────────────────────────────────────
    {
        'bagian': 'SECOND_PROSES',
        'file':   'data_nama_karu_asstn_departemen_second_proses.xlsx',
        'sheets': [
            {
                'type':         'jabatan',
                'sheet':        'Sheet1',
                'header_row':   6,           # baris header: NO | REGU ROHALI | ... | JABATAN
                'col_no':       1,
                'cols_regu':    [2, 3, 4],   # kolom per regu
                'col_jabatan':  5,           # kolom JABATAN (1-based)
                'karu_prefix':  'REGU ',
                'karu_jabatan': ['KARU'],                    # nilai di kolom JABATAN → jadi karu
                'asisten_jabatan': ['ASSTN KARU', 'ASISTEN'], # nilai → jadi asisten
            },
        ],
    },

]

# ═══════════════════════════════════════════════════════════════════


def resolve_sheet(file_path, sheet_name):
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    if sheet_name in sheets:
        return sheet_name, sheets
    for s in sheets:
        if s.strip().lower() == sheet_name.strip().lower():
            print(f"  ℹ️  Sheet '{sheet_name}' → pakai '{s}' (nama mirip)")
            return s, sheets
    return None, sheets


def normalize_columns(df):
    import pandas as pd
    df.columns = [str(c).replace('\n', ' ').strip() for c in df.columns]
    return df


def auto_detect_header(file_path, sheet_name, target_cols, max_scan=5):
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[sheet_name]
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=True)):
        cells = [str(c).replace('\n', ' ').strip() if c is not None else '' for c in row]
        if all(col in cells for col in target_cols):
            wb.close()
            return row_idx
    wb.close()
    return 0


def read_produk(file_path, sheet_cfg, bagian):
    import pandas as pd
    sheet_raw = sheet_cfg['sheet']
    sheet, all_sheets = resolve_sheet(file_path, sheet_raw)
    if sheet is None:
        print(f"  ⚠️  Sheet '{sheet_raw}' tidak ditemukan. Sheet tersedia: {all_sheets}")
        return []

    header_r = sheet_cfg.get('header', None)
    if header_r is None:
        header_r = auto_detect_header(file_path, sheet, ['KODE BARANG', 'NAMA PRODUK'])
        if header_r > 0:
            print(f"  ℹ️  Header ditemukan di baris Excel {header_r + 1}")

    try:
        df = pd.read_excel(file_path, sheet_name=sheet, header=header_r)
        df = normalize_columns(df)
        df = df.loc[:, ~df.columns.duplicated()]
    except Exception as e:
        print(f"  ⚠️  Gagal baca sheet '{sheet}': {e}")
        return []

    results = []
    cols = list(df.columns)
    for i in range(len(cols)):
        col_kode = cols[i]
        if "KODE" in col_kode.upper():
            if i + 1 < len(cols):
                col_nama = cols[i + 1]
                if "PRODUK" in col_nama.upper():
                    temp = df[[col_kode, col_nama]].copy()
                    temp.columns = ['kode', 'nama']
                    temp = temp.dropna()
                    temp['kode'] = temp['kode'].astype(str).str.strip()
                    temp['nama'] = temp['nama'].astype(str).str.strip()
                    temp = temp[
                        (temp['kode'] != '') &
                        (temp['nama'] != '') &
                        (temp['kode'].str.lower() != 'nan') &
                        (temp['nama'].str.lower() != 'nan')
                    ]
                    temp['bagian'] = bagian
                    results.extend(temp.to_dict('records'))

    unique = {}
    for r in results:
        unique[(r['kode'], r['nama'])] = r
    results = list(unique.values())
    print(f"  ✔  Multi-table detect: {len(results)} produk")
    return results


def read_mesin(file_path, sheet_cfg, bagian):
    import pandas as pd
    sheet_raw = sheet_cfg['sheet']
    col_m     = sheet_cfg['col_mesin']
    sheet, all_sheets = resolve_sheet(file_path, sheet_raw)
    if sheet is None:
        print(f"  ⚠️  Sheet '{sheet_raw}' tidak ditemukan untuk mesin.")
        return []

    header_r = sheet_cfg.get('header', None)
    if header_r is None:
        header_r = auto_detect_header(file_path, sheet, [col_m])

    try:
        df = pd.read_excel(file_path, sheet_name=sheet, header=header_r)
        df = normalize_columns(df)
    except Exception as e:
        print(f"  ⚠️  Gagal baca sheet '{sheet}': {e}")
        return []

    if col_m not in df.columns:
        print(f"  ⚠️  Kolom mesin '{col_m}' tidak ada. Tersedia: {list(df.columns)}")
        return []

    mesins = df[col_m].dropna().astype(str).str.strip().unique().tolist()
    mesins = [m for m in mesins if m and m.lower() not in ('nan', col_m.lower())]
    return sorted(mesins)


# ─── Parser karyawan per TYPE ─────────────────────────────────────

def read_karyawan_inject(file_path, sheet_cfg):
    """TYPE A: Header horizontal, kolom per Karu."""
    import openpyxl
    sheet_name  = sheet_cfg['sheet']
    header_row  = sheet_cfg['header_row']
    col_mesin   = sheet_cfg['col_mesin']
    cols_as     = sheet_cfg['cols_asisten']
    karu_pfx    = sheet_cfg.get('karu_prefix', 'Karu ')

    sheet, all_sheets = resolve_sheet(file_path, sheet_name)
    if sheet is None:
        print(f"  ⚠️  Sheet '{sheet_name}' tidak ditemukan. Tersedia: {all_sheets}")
        return set(), set()

    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[sheet]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Ambil nama Karu dari header
    karu_names = {}
    header_cells = all_rows[header_row - 1]
    for col_idx in cols_as:
        raw = header_cells[col_idx - 1]
        if raw:
            raw_str   = str(raw).strip()
            nama_karu = (raw_str[len(karu_pfx):].strip()
                         if raw_str.lower().startswith(karu_pfx.lower())
                         else raw_str)
            karu_names[col_idx] = nama_karu

    karu_set    = set(karu_names.values())
    asisten_set = set()

    for row in all_rows[header_row:]:
        mesin_val = row[col_mesin - 1]
        if not mesin_val:
            continue
        for col_idx in cols_as:
            as_val = row[col_idx - 1]
            if not as_val:
                continue
            asisten = str(as_val).strip()
            if asisten and asisten.lower() not in ('none', ''):
                asisten_set.add(asisten)

    return karu_set, asisten_set


def read_karyawan_pet(file_path, sheet_cfg):
    """
    TYPE B: Kolom vertikal — Col A=Karu, Col B=Asisten, Col C=No Mesin.
    Karu yang pakai formula (=+A4) diteruskan dari baris sebelumnya.
    """
    import openpyxl
    sheet_name  = sheet_cfg['sheet']
    header_row  = sheet_cfg['header_row']

    sheet, all_sheets = resolve_sheet(file_path, sheet_name)
    if sheet is None:
        print(f"  ⚠️  Sheet '{sheet_name}' tidak ditemukan. Tersedia: {all_sheets}")
        return set(), set()

    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[sheet]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    karu_set    = set()
    asisten_set = set()
    current_karu = None

    # Data mulai setelah header_row (1-based)
    for row in all_rows[header_row:]:
        col_a = row[0]   # Karu (bisa formula string → skip)
        col_b = row[1]   # Asisten
        # col_c = row[2] # No Mesin (tidak dipakai untuk karyawan)

        # Kolom A: ambil nama karu hanya jika bukan formula dan bukan None
        if col_a and not str(col_a).startswith('='):
            val = str(col_a).strip()
            # Skip jika ini header yang tidak sengaja terbaca
            if val.upper() not in ('KETUA REGU', 'NAMA ASISTEN', 'NO MC', ''):
                current_karu = val
                karu_set.add(current_karu)

        # Kolom B: nama asisten
        if col_b:
            val = str(col_b).strip()
            skip_vals = ('nama asisten', 'no mc', 'ketua regu', 'none', '')
            if val.lower() not in skip_vals:
                asisten_set.add(val)

    return karu_set, asisten_set


def read_karyawan_regu(file_path, sheet_cfg):
    """
    TYPE C: Header baris N, kolom per regu.
    Nama karu = header kolom (setelah hapus prefix).
    Semua isi kolom = asisten.
    """
    import openpyxl
    sheet_name  = sheet_cfg['sheet']
    header_row  = sheet_cfg['header_row']
    cols_regu   = sheet_cfg['cols_regu']
    karu_pfx    = sheet_cfg.get('karu_prefix', 'REGU ')

    sheet, all_sheets = resolve_sheet(file_path, sheet_name)
    if sheet is None:
        print(f"  ⚠️  Sheet '{sheet_name}' tidak ditemukan. Tersedia: {all_sheets}")
        return set(), set()

    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[sheet]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_cells = all_rows[header_row - 1]

    # Ambil nama karu dari header kolom
    karu_set    = set()
    karu_names  = {}   # col_idx (1-based) → nama karu
    for col_idx in cols_regu:
        raw = header_cells[col_idx - 1]
        if raw:
            raw_str = str(raw).strip()
            nama    = (raw_str[len(karu_pfx):].strip()
                       if karu_pfx and raw_str.upper().startswith(karu_pfx.upper())
                       else raw_str)
            karu_names[col_idx] = nama
            karu_set.add(nama)

    # Ambil asisten dari baris data
    asisten_set = set()
    for row in all_rows[header_row:]:   # baris setelah header
        for col_idx in cols_regu:
            val = row[col_idx - 1]
            if not val:
                continue
            nama = str(val).strip()
            if nama and nama.lower() not in ('none', ''):
                asisten_set.add(nama)

    return karu_set, asisten_set


def read_karyawan_jabatan(file_path, sheet_cfg):
    """
    TYPE D: Header baris N, kolom per regu + kolom JABATAN.
    Baris dengan JABATAN=KARU → nama jadi karu.
    Baris lain → nama jadi asisten.
    """
    import openpyxl
    sheet_name     = sheet_cfg['sheet']
    header_row     = sheet_cfg['header_row']
    cols_regu      = sheet_cfg['cols_regu']
    col_jabatan    = sheet_cfg['col_jabatan']
    karu_pfx       = sheet_cfg.get('karu_prefix', 'REGU ')
    karu_jab       = [j.upper() for j in sheet_cfg.get('karu_jabatan', ['KARU'])]
    asisten_jab    = [j.upper() for j in sheet_cfg.get('asisten_jabatan', ['ASSTN KARU', 'ASISTEN'])]

    sheet, all_sheets = resolve_sheet(file_path, sheet_name)
    if sheet is None:
        print(f"  ⚠️  Sheet '{sheet_name}' tidak ditemukan. Tersedia: {all_sheets}")
        return set(), set()

    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[sheet]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_cells = all_rows[header_row - 1]

    # Nama karu bisa dari header ATAU dari baris data dengan jabatan KARU
    karu_set    = set()
    asisten_set = set()

    # Opsional: nama di header sebagai referensi saja (tidak dipakai jika ada baris KARU)
    for row in all_rows[header_row:]:
        jabatan_val = row[col_jabatan - 1]
        if not jabatan_val:
            continue
        jabatan = str(jabatan_val).strip().upper()

        for col_idx in cols_regu:
            val = row[col_idx - 1]
            if not val:
                continue
            nama = str(val).strip()
            if not nama or nama.lower() in ('none', ''):
                continue
            # Title case untuk konsistensi
            if jabatan in karu_jab:
                karu_set.add(nama.title())
            elif jabatan in asisten_jab:
                asisten_set.add(nama.title())

    return karu_set, asisten_set


def read_karyawan(file_path, sheet_cfg, bagian):
    """Dispatcher: pilih parser sesuai 'type' di konfigurasi."""
    t = sheet_cfg.get('type', 'inject')
    if t == 'inject':
        return read_karyawan_inject(file_path, sheet_cfg)
    elif t == 'pet':
        return read_karyawan_pet(file_path, sheet_cfg)
    elif t == 'regu':
        return read_karyawan_regu(file_path, sheet_cfg)
    elif t == 'jabatan':
        return read_karyawan_jabatan(file_path, sheet_cfg)
    else:
        print(f"  ⚠️  Tipe sheet tidak dikenal: '{t}'")
        return set(), set()


# ─── Upload functions ─────────────────────────────────────────────

def upload_karyawan(db, all_karyawan_by_bagian, dry_run=False):
    total = sum(len(v['karu']) + len(v['asisten']) for v in all_karyawan_by_bagian.values())
    print(f"\n👤 Mengupload {total} karyawan ke 'master_karyawan'...")

    if not dry_run:
        batch = db.batch()

    count = 0
    for bagian, data in all_karyawan_by_bagian.items():
        karu_list    = sorted(data['karu'])
        asisten_list = sorted(data['asisten'])
        print(f"  [{bagian}] {len(karu_list)} karu   : {', '.join(karu_list)}")
        print(f"  [{bagian}] {len(asisten_list)} asisten: "
              f"{', '.join(asisten_list[:6])}{'...' if len(asisten_list) > 6 else ''}")

        for nama in karu_list:
            safe   = nama.replace(' ', '_').replace('/', '_')
            doc_id = f"{bagian}_karu_{safe}"
            if dry_run:
                if count < 3:
                    print(f"  [DRY-RUN] → karu: {nama} ({bagian})")
            else:
                ref = db.collection('master_karyawan').document(doc_id)
                batch.set(ref, {'nama': nama, 'bagian': bagian, 'role': 'karu'})
            count += 1

        for nama in asisten_list:
            safe   = nama.replace(' ', '_').replace('/', '_')
            doc_id = f"{bagian}_asisten_{safe}"
            if dry_run:
                if count < 6:
                    print(f"  [DRY-RUN] → asisten: {nama} ({bagian})")
            else:
                ref = db.collection('master_karyawan').document(doc_id)
                batch.set(ref, {'nama': nama, 'bagian': bagian, 'role': 'asisten'})
            count += 1

    if not dry_run:
        batch.commit()

    print(f"  ✅ {count} karyawan {'akan di' if dry_run else ''}upload ke 'master_karyawan'")
    return count


def upload_produk(db, all_produk, dry_run=False):
    print(f"\n📦 Mengupload {len(all_produk)} produk ke 'master_produk'...")

    kode_count = {}
    batch      = db.batch() if not dry_run else None
    count      = 0
    BATCH_SIZE = 400

    for row in all_produk:
        kode   = row['kode']
        nama   = row['nama']
        bagian = row['bagian']

        kode_count[kode] = kode_count.get(kode, 0) + 1
        doc_id = kode if kode_count[kode] == 1 else f"{kode}_{kode_count[kode]}"

        if dry_run:
            if count < 5:
                print(f"  [DRY-RUN] → {doc_id}: {nama[:50]} ({bagian})")
        else:
            ref = db.collection('master_produk').document(doc_id)
            batch.set(ref, {'kode': kode, 'nama': nama, 'bagian': bagian, 'docId': doc_id})

        count += 1
        if not dry_run and count % BATCH_SIZE == 0:
            batch.commit()
            batch = db.batch()
            print(f"  → {count} produk terupload...")

    if not dry_run:
        batch.commit()

    print(f"  ✅ {count} produk {'akan di' if dry_run else ''}upload ke 'master_produk'")
    return count


def upload_mesin(db, all_mesin_by_bagian, dry_run=False):
    total = sum(len(v) for v in all_mesin_by_bagian.values())
    print(f"\n⚙️  Mengupload {total} mesin ke 'master_mesin'...")

    if not dry_run:
        batch = db.batch()

    for bagian, mesins in all_mesin_by_bagian.items():
        print(f"  [{bagian}] {len(mesins)} mesin: "
              f"{', '.join(mesins[:8])}{'...' if len(mesins) > 8 else ''}")
        for m in mesins:
            doc_id = f"{bagian}_{m.replace(' ', '_').replace('-', '').replace('_', '')}"
            if not dry_run:
                ref = db.collection('master_mesin').document(doc_id)
                batch.set(ref, {'noMesin': m, 'bagian': bagian})

    if not dry_run:
        batch.commit()

    print(f"  ✅ {total} mesin {'akan di' if dry_run else ''}upload ke 'master_mesin'")
    return total


def clear_collection(db, col_name):
    print(f"  🗑️  Menghapus koleksi '{col_name}'...")
    col_ref = db.collection(col_name)
    deleted = 0
    while True:
        docs  = col_ref.limit(400).stream()
        batch = db.batch()
        n     = 0
        for doc in docs:
            batch.delete(doc.reference)
            n += 1
        if n == 0:
            break
        batch.commit()
        deleted += n
        print(f"     → {deleted} dokumen dihapus...")
    print(f"  ✅ '{col_name}' bersih ({deleted} dokumen dihapus)")


# ─── Main ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Upload master data ke Firestore')
    parser.add_argument('sa_key',       help='Path ke serviceAccountKey.json')
    parser.add_argument('--file',       help='Upload satu file tertentu saja')
    parser.add_argument('--bagian',     help='Bagian untuk --file (misal: INJECT)')
    parser.add_argument('--clear',      action='store_true', help='Hapus data lama sebelum upload')
    parser.add_argument('--dry-run',    action='store_true', help='Simulasi, tidak upload')
    args = parser.parse_args()

    if not os.path.exists(args.sa_key):
        print(f"❌ File serviceAccountKey.json tidak ditemukan: '{args.sa_key}'")
        sys.exit(1)

    if args.file:
        if not args.bagian:
            print("❌ Jika pakai --file, harus juga pakai --bagian (misal: --bagian PET)")
            sys.exit(1)
        db_list       = [d for d in DATABASES          if d['bagian'] == args.bagian]
        karyawan_list = [d for d in KARYAWAN_DATABASES if d['bagian'] == args.bagian]
        if not db_list and not karyawan_list:
            print(f"⚠️  Bagian '{args.bagian}' tidak ditemukan di konfigurasi.")
            sys.exit(1)
    else:
        db_list       = DATABASES
        karyawan_list = KARYAWAN_DATABASES

    print("=" * 60)
    print("🚀 UPLOAD MASTER DATA KE FIRESTORE")
    if args.dry_run:
        print("   [MODE DRY-RUN — tidak ada yang diupload]")
    print("=" * 60)

    if not args.dry_run:
        cred  = credentials.Certificate(args.sa_key)
        firebase_admin.initialize_app(cred)
        db_fs = firestore.client()
        print("✅ Firebase terhubung\n")
    else:
        db_fs = None
        print("✅ Mode dry-run aktif\n")

    if args.clear and not args.dry_run:
        print("🗑️  Menghapus data lama...")
        clear_collection(db_fs, 'master_produk')
        clear_collection(db_fs, 'master_mesin')
        clear_collection(db_fs, 'master_karyawan')
        print()

    # ── Produk & Mesin ────────────────────────────────────────────
    all_produk          = []
    all_mesin_by_bagian = {}

    for cfg in db_list:
        bagian    = cfg['bagian']
        file_path = cfg['file']
        print(f"📂 [{bagian}] Membaca: {file_path}")

        if not os.path.exists(file_path):
            print(f"  ⚠️  File tidak ditemukan, dilewati: '{file_path}'\n")
            continue

        produk_rows = []
        for sc in cfg.get('produk_sheets', []):
            rows = read_produk(file_path, sc, bagian)
            produk_rows.extend(rows)
            print(f"  ✔  Sheet '{sc['sheet']}': {len(rows)} produk")

        seen = set()
        unique_p = []
        for r in produk_rows:
            key = (r['kode'], r['nama'])
            if key not in seen:
                seen.add(key)
                unique_p.append(r)
        all_produk.extend(unique_p)
        print(f"  📦 Total produk [{bagian}]: {len(unique_p)} (unik)")

        mesins     = []
        seen_mesin = set()
        for sc in cfg.get('mesin_sheets', []):
            for m in read_mesin(file_path, sc, bagian):
                if m not in seen_mesin:
                    seen_mesin.add(m)
                    mesins.append(m)
        all_mesin_by_bagian[bagian] = sorted(mesins)
        print(f"  ⚙️  Total mesin  [{bagian}]: {len(mesins)}\n")

    # ── Karyawan ──────────────────────────────────────────────────
    all_karyawan_by_bagian = {}

    for cfg in karyawan_list:
        bagian    = cfg['bagian']
        file_path = cfg['file']
        print(f"📂 [{bagian}] Membaca karyawan: {file_path}")

        if not os.path.exists(file_path):
            print(f"  ⚠️  File tidak ditemukan, dilewati: '{file_path}'\n")
            continue

        all_karu    = set()
        all_asisten = set()

        for sc in cfg['sheets']:
            print(f"  📋 Sheet: '{sc['sheet']}' (type={sc.get('type','inject')})")
            karu_set, asisten_set = read_karyawan(file_path, sc, bagian)
            all_karu    |= karu_set
            all_asisten |= asisten_set

        # Overlap: nama yang sama di karu & asisten → simpan sebagai karu
        overlap = all_karu & all_asisten
        if overlap:
            print(f"  ⚠️  Overlap (disimpan sbg karu): {overlap}")
            all_asisten -= overlap

        all_karyawan_by_bagian[bagian] = {'karu': all_karu, 'asisten': all_asisten}
        print(f"  👤 Total [{bagian}]: {len(all_karu)} karu, {len(all_asisten)} asisten\n")

    # ── Upload ────────────────────────────────────────────────────
    total_produk   = upload_produk(db_fs, all_produk,              dry_run=args.dry_run)
    total_mesin    = upload_mesin(db_fs, all_mesin_by_bagian,      dry_run=args.dry_run)
    total_karyawan = upload_karyawan(db_fs, all_karyawan_by_bagian, dry_run=args.dry_run)

    # ── Ringkasan ─────────────────────────────────────────────────
    print(f"""
{'=' * 60}
✅ SELESAI! {'[DRY-RUN] ' if args.dry_run else ''}Data tersimpan di Firestore:
   • master_produk    : {total_produk} dokumen
   • master_mesin     : {total_mesin} dokumen
   • master_karyawan  : {total_karyawan} dokumen

   Breakdown per bagian:""")

    bagian_counts = {}
    for r in all_produk:
        bagian_counts[r['bagian']] = bagian_counts.get(r['bagian'], 0) + 1
    all_bagian = set(list(bagian_counts) + list(all_karyawan_by_bagian))
    for b in sorted(all_bagian):
        p = bagian_counts.get(b, 0)
        m = len(all_mesin_by_bagian.get(b, []))
        k = all_karyawan_by_bagian.get(b, {})
        print(f"   ─ {b:<14}: {p} produk, {m} mesin, "
              f"{len(k.get('karu',[]))} karu, {len(k.get('asisten',[]))} asisten")

    print(f"\nAplikasi React Native akan otomatis fetch data terbaru.\n{'=' * 60}")


if __name__ == '__main__':
    main()